#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Dorian Wang 2017.09.13
# lesson: get all detail information from several excel file

import os
import openpyxl
import win32com.client

detail = []

#  loop all file, save xls to xlsx and remove old file
print('==========Check xls file==========')
for folder_name, subfolders, filenames in os.walk('./Excel/'):
    for filename in filenames:
        report = os.path.join(os.getcwd() + '\\Excel\\' + filename)
        # print(report)
        if report[-3:] == 'xls':
            excel = win32com.client.gencache.EnsureDispatch('Excel.Application')  # 要看MIME手册
            wb = excel.Workbooks.Open(report)
            wb.SaveAs(report + 'x', FileFormat=51)
            wb.Close()
            excel.Application.Quit()
            os.remove(report)

#  loop all excel file, get detail data
print('==========Begin Read Data==========')
for folder_name, subfolders, filenames in os.walk('./Excel/'):
    for filename in filenames:
        print(filename)
        workbook = openpyxl.load_workbook('./Excel/' + filename, data_only=True)
        sheet_names = workbook.get_sheet_names()
        # print(sheet_names)
        for sheetname in sheet_names:
            print(sheetname)
            worksheet = workbook.get_sheet_by_name(sheetname)
            # print(worksheet.sheet_state)
            if worksheet.sheet_state == 'visible':
                sheet_end = False
                build_no = worksheet['B2'].value
                build_level = worksheet['D2'].value
                # print(worksheet.max_row)
                for line in range(4, worksheet.max_row + 1):
                    serial_num = worksheet['A' + str(line)].value
                    # print(line, serial_num)
                    # print(type(serial_num))
                    if serial_num == None:
                        sheet_end = True
                        break
                    elif (serial_num == '序号') or (serial_num == '楼幢号：'):
                        continue
                    # print(serial_num)
                    room_num = worksheet['B' + str(line)].value

                    size_data = worksheet['C' + str(line)].value
                    # print(size_data)
                    # print(type(size_data))
                    if (size_data == None) or (size_data == '') :
                        break
                    elif str(size_data)[-1:] == '㎡':
                        room_size = int(float(str(size_data)[0:-1]))
                    else:
                        room_size = int(float(size_data))

                    price_data = worksheet['D' + str(line)].value
                    # print(price_data)
                    if price_data == None:
                        break
                    elif str(price_data)[-1:] == '元':
                        room_price = round(float(str(price_data)[0:-1]) / 10000, 2)
                    elif str(price_data)[-4:] == '（毛坯）':
                        room_price = round(float(str(price_data)[0:-4]) / 10000, 2)
                    else:
                        room_price = round(float(price_data) / 10000, 2)

                    value_data = worksheet['E' + str(line)].value
                    # print(value_data)
                    if value_data == None:
                        break
                    elif str(value_data)[-1:] == '元':
                        room_value = int(float(str(value_data)[0:-1]) / 10000)
                    else:
                        room_value = int(float(value_data) / 10000)

                    detail.append({'filename': filename, 'sheetname': sheetname, 'build_no': build_no, 'build_level': build_level, 'serial_num': serial_num, 'room_num': room_num, 'room_size': room_size, 'room_price': room_price,
                                   'room_value': room_value})

for line in detail:
    # print(line)
    print(line['filename'], line['sheetname'], line['build_no'], line['build_level'], line['serial_num'], line['room_num'], line['room_size'], line['room_price'], line['room_value'])

#  save detail data into new sum excel file
print('==========Write Data to Sum Excel==========')
sum_workbook = openpyxl.Workbook()
sum_sheet = sum_workbook.active
sum_sheet['A1'] = '来源文件'
sum_sheet['B1'] = 'Sheet页'
sum_sheet['C1'] = '楼幢号'
sum_sheet['D1'] = '楼高'
sum_sheet['E1'] = '序号'
sum_sheet['F1'] = '房号'
sum_sheet['G1'] = '面积'
sum_sheet['H1'] = '单价（万）'
sum_sheet['I1'] = '总价（万）'
i = 1
for line in detail:
    i = i + 1
    sum_sheet['A' + str(i)] = line['filename']
    sum_sheet['B' + str(i)] = line['sheetname']
    sum_sheet['C' + str(i)] = line['build_no']
    sum_sheet['D' + str(i)] = line['build_level']
    sum_sheet['E' + str(i)] = line['serial_num']
    sum_sheet['F' + str(i)] = line['room_num']
    sum_sheet['G' + str(i)] = line['room_size']
    sum_sheet['H' + str(i)] = line['room_price']
    sum_sheet['I' + str(i)] = line['room_value']

sum_workbook.save('西安房价汇总2018.xlsx')
