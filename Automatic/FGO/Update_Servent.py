#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Dorian Wang 2017.09.13
# get all detail_list information from servant excel file, then update Dorian.fgh for FGO Helper.

import openpyxl

detail_list = []

print('==========Read Data from FGO excel==========')

filename_excel = 'd:\Dorian\Documents\Life\Mind and Read\Read Game Note\FGO\FGO.xlsx'
workbook = openpyxl.load_workbook(filename_excel, data_only=True)
worksheet = workbook.get_sheet_by_name('Table')

# print(worksheet.max_row)
for line in range(2, worksheet.max_row + 1):
    servant_no = worksheet['A' + str(line)].value
    ascension = worksheet['F' + str(line)].value
    skill1 = worksheet['L' + str(line)].value
    skill2 = worksheet['M' + str(line)].value
    skill3 = worksheet['N' + str(line)].value
    priority = worksheet['S' + str(line)].value
    if ascension is None:
        ascension = 0
    if skill1 is None:
        skill1 = 1
    if skill2 is None:
        skill2 = 1
    if skill3 is None:
        skill3 = 1
    if str(priority) <= str(5):
        detail_list.append({'servant_no': servant_no, 'ascension': ascension, 'skill1': skill1, 'skill2': skill2, 'skill3': skill3, 'priority': priority})

for line in detail_list:
    print(line['servant_no'], line['ascension'], line['skill1'], line['skill2'], line['skill3'], line['priority'])

# print(detail_list.index({'servant_no': 109, 'ascension': 1, 'skill1': 1, 'skill2': 1, 'skill3': 1, 'priority': 5}))


#  save detail_list data into FGO Helper file
print('==========Compare Data with FGO Helper==========')
filename_FGOHelper = 'd:\Dorian\Documents\Life\Mind and Read\Read Game Note\FGO\Dorian.fgh'
old_file = open(filename_FGOHelper, 'rb')
eachlines = old_file.readlines()
old_file.close()
# print(eachlines)

results = []
for line in eachlines[1:]:
    # print(line.decode('utf-8'))
    data = line.split()
    ID = data[0]
    for servant in detail_list:
        update = False
        if int(ID) + 1 == int(servant['servant_no']):
            # print(data)
            if int(data[2]) != int(servant['ascension']):
                # print(data[2])
                print('servant_no:', servant['servant_no'], 'ascension:', servant['ascension'])
                data[2] = servant['ascension']
                update = True
            if int(data[4]) != int(servant['skill1']):
                # print(data[4])
                print('servant_no:', servant['servant_no'], 'skill1:', servant['skill1'])
                data[4] = servant['skill1']
                update = True
            if int(data[5]) != int(servant['skill2']):
                # print(data[5])
                print('servant_no:', servant['servant_no'], 'skill2:', servant['skill2'])
                data[5] = servant['skill2']
                update = True
            if int(data[6]) != int(servant['skill3']):
                # print(data[6])
                print('servant_no:', servant['servant_no'], 'skill3:', servant['skill3'])
                data[6] = servant['skill3']
                update = True
        if update:
            # print('Before Updated:')
            # print(line)
            # print('After Updated:')
            # # line = '\\t'.join(data)
            # print(data)
            print()
            break

# print(results)
# new_file = open(filename_FGOHelper, 'w')
# new_file.writelines(results)
# new_file.close()

print('==========Done.==========')
