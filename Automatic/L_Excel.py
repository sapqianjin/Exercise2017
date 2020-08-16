#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Dorian Wang 2017.09.13
# lesson: excel file processing

import openpyxl
detail = []

filename = './Apartment/11、中海•曲江大城 11#、12#、13#、14#.xlsx'
workbook = openpyxl.load_workbook(filename)
# print(type(workbook))
sheet_names = workbook.get_sheet_names()
# print(sheet_names)
for sheetname in sheet_names:
    worksheet = workbook.get_sheet_by_name(sheetname)
    build_no = worksheet['B2'].value
    build_level = worksheet['D2'].value

    for line in range(4, worksheet.max_row+1):
        serial_num = worksheet['A' + str(line)].value
        room_num = worksheet['B' + str(line)].value
        room_size = int(worksheet['C' + str(line)].value)
        room_price = round(worksheet['D' + str(line)].value / 10000, 2)
        room_value = int(worksheet['E' + str(line)].value / 10000)

        detail.append({'filename':filename, 'sheetname':sheetname,'build_no':build_no, 'build_level':build_level, 'serial_num':serial_num, 'room_num':room_num, 'room_size':room_size, 'room_price':room_price, 'room_amount':room_value})

for line in detail:
    # print(line)
    print(line['filename'], line['sheetname'],line['build_no'],line['build_level'], line['serial_num'], line['room_num'],line['room_size'],line['room_price'],line['room_amount'])

# 楼幢号B2
# 楼高D2
# 序号A4到某个为空的Ax
# 房号从B4
# 面积C4
# 单价D4 / 10000
# 总价D4 / 10000

# wb = openpyxl.Workbook()
# ws = wb.active
# ws['A1'] = 'test'
# wb.save('Test.xlsx')